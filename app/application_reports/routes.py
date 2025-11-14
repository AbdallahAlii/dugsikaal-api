# #
# # # app/application_reports/routes.py
#
# from __future__ import annotations
# import logging
# import csv
# import io
# from typing import Dict, Any
# from datetime import datetime, date
#
# from flask import Blueprint, request, jsonify, current_app, Response, g, make_response
# from werkzeug.exceptions import Forbidden
#
# from app.application_reports.core.engine import create_report_engine
# from app.application_reports.bootstrap import bootstrap_reports
# from app.application_reports.core.cache import ReportCache
# from app.application_reports.security import validate_report_access
# from app.auth.deps import get_current_user
# from app.security.rbac_effective import AffiliationContext
# from config.database import db
#
# bp = Blueprint('reports', __name__, url_prefix='/api/reports')
# log = logging.getLogger(__name__)
#
# # Reports we treat as "hot" (live by default; rarely cached)
# _HOT_REPORTS = {
#     "General Ledger",
#     "Accounts Payable Summary",
#     "Accounts Receivable Summary",
#     "Accounts Payable",
#     "Accounts Receivable",
#     "Stock Ledger",
#     "Trial Balance",
#     "Balance Sheet",
#     "Profit and Loss",
# }
#
#
# # ---- auth ctx helper ---------------------------------------------------------
# def _ctx() -> AffiliationContext:
#     _ = get_current_user()
#     ctx: AffiliationContext = getattr(g, "auth", None)
#     if not ctx:
#         raise PermissionError("Authentication context not found.")
#     return ctx
#
#
# # ---- single engine & cache ---------------------------------------------------
# _engine_instance = None
# _report_cache: ReportCache | None = None
#
#
# def _get_engine():
#     global _engine_instance
#     if _engine_instance is None:
#         _engine_instance = create_report_engine(db.session)
#         bootstrap_reports(_engine_instance)
#         log.info("✅ Reports engine initialized and bootstrapped")
#     return _engine_instance
#
#
# def _get_cache() -> ReportCache:
#     global _report_cache
#     if _report_cache is None:
#         enabled = current_app.config.get('REPORT_CACHE_ENABLED', True)
#         ttl = int(current_app.config.get('REPORT_CACHE_TTL', 30))
#         _report_cache = ReportCache(enabled=enabled, default_ttl=ttl)
#         log.info("🧠 Report cache initialized (enabled=%s, ttl=%s)", enabled, ttl)
#     return _report_cache
#
#
# def _to_int_or_none(v):
#     try:
#         return int(v) if v is not None and str(v).strip() != "" else None
#     except Exception:
#         return None
#
#
# # ---- date helpers (ERP-style) -----------------------------------------------
# # Display format everywhere in API responses (filters, not data rows):
# #   MM/DD/YYYY  (e.g., 10/20/2025)
# _DISPLAY_FMT = "%d-%m-%Y"
#
# # Accept these inputs for inbound filters:
# # - With slashes:  MM/DD/YYYY (primary), DD/MM/YYYY, YYYY/MM/DD
# # - With dashes:   YYYY-MM-DD (primary), DD-MM-YYYY, MM-DD-YYYY
# # - ISO datetime:  take date part
# _DATE_PARSE_PATTERNS = [
#     "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d",
#     "%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y",
# ]
#
#
# def _parse_date_flex(s: str) -> date | None:
#     s = (s or "").strip()
#     if not s:
#         return None
#     # If full ISO datetime is sent, slice date part first
#     if "T" in s or " " in s:
#         try:
#             dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
#             return dt.date()
#         except Exception:
#             pass
#     # Try the patterns in order
#     for fmt in _DATE_PARSE_PATTERNS:
#         try:
#             return datetime.strptime(s, fmt).date()
#         except Exception:
#             continue
#     # Final attempt: raw ISO date
#     try:
#         return datetime.fromisoformat(s).date()
#     except Exception:
#         return None
#
#
# def _format_date_out(d: date | datetime | None) -> str | None:
#     if d is None:
#         return None
#     if isinstance(d, datetime):
#         d = d.date()
#     return d.strftime(_DISPLAY_FMT)
#
#
# def _format_filter_dates(filters: Dict[str, Any]) -> Dict[str, Any]:
#     """
#     Return a shallow copy of filters with *_date style keys formatted
#     as MM/DD/YYYY strings and WITHOUT time.
#     """
#     if not isinstance(filters, dict):
#         return filters
#     out = dict(filters)
#     date_keys = set(k for k in out.keys()
#                     if k.endswith("_date") or k in {"date", "as_on_date", "period_date", "report_date"})
#     for k in date_keys:
#         v = out.get(k)
#         if isinstance(v, (date, datetime)):
#             out[k] = _format_date_out(v)
#         elif isinstance(v, str):
#             parsed = _parse_date_flex(v)
#             out[k] = _format_date_out(parsed) if parsed else v
#     return out
#
#
# def _booly(v) -> bool | None:
#     if isinstance(v, bool):
#         return v
#     if isinstance(v, str):
#         return v.lower() in ("true", "1", "yes", "y", "on")
#     return None
#
#
# def _should_cache(report_name: str, report_meta, filters: Dict[str, Any]) -> tuple[bool, int]:
#     """
#     Decide if we should cache, and for how long.
#     Priority:
#       1) request override: nocache=1 or cache=false -> no cache
#       2) report meta cache_enabled True/False
#       3) hot reports default False, others True
#     TTL:
#       - report meta cache_ttl_s if set, else app config
#     """
#     client_nocache = _booly(filters.get('nocache'))
#     client_cache_flag = _booly(filters.get('cache'))
#     if client_nocache or (client_cache_flag is False):
#         return (False, 0)
#
#     if report_meta and report_meta.cache_enabled is not None:
#         enabled = bool(report_meta.cache_enabled)
#     else:
#         enabled = (report_name not in _HOT_REPORTS)
#
#     ttl = int(report_meta.cache_ttl_s or current_app.config.get('REPORT_CACHE_TTL', 30))
#     return (enabled, ttl)
#
#
# # ---- routes ------------------------------------------------------------------
# @bp.route('/<report_name>', methods=['GET', 'POST'])
# def execute_report(report_name: str):
#     """
#     Execute any report by name with strict tenant security.
#     Requires `company` in filters for data-bearing reports.
#     """
#     try:
#         context = _ctx()
#         engine = _get_engine()
#
#         # Parse filters
#         filters = parse_filters(request)
#         log.debug(f"🔧 [ROUTE] Initial filters for {report_name}: {filters}")
#
#         # Resolve and validate scope BEFORE cache/execute
#         company_id = _to_int_or_none(filters.get('company'))
#         branch_id = _to_int_or_none(filters.get('branch_id'))  # optional
#         validate_report_access(context=context, company_id=company_id, branch_id=branch_id)
#
#         # Cache decision
#         report_meta = engine.get_report_meta(report_name)
#         use_cache, ttl = _should_cache(report_name, report_meta, filters)
#
#         cache = _get_cache()
#         if use_cache:
#             cached_result = cache.get(report_name, filters)
#             if cached_result:
#                 cached_result = {**cached_result, "filters": _format_filter_dates(cached_result.get("filters", {}))}
#                 resp = make_response(jsonify({
#                     "success": True,
#                     "cached": True,
#                     **cached_result
#                 }))
#                 resp.headers["Cache-Control"] = f"public, max-age={ttl}"
#                 return resp
#
#         # Execute
#         log.debug(f"🚀 [ROUTE] Executing report {report_name} with filters: {filters}")
#         result = engine.execute_report(report_name, filters, context)
#         log.debug(f"✅ [ROUTE] Report executed successfully. Data rows: {len(result.get('data', []))}")
#
#         # Debug: Check date formats in data
#         if result.get('data'):
#             sample_row = result['data'][0]
#             posting_date = sample_row.get('posting_date')
#             due_date = sample_row.get('due_date')
#             log.debug(f"📅 [ROUTE] Sample row dates - posting_date type: {type(posting_date)}, value: {posting_date}")
#             log.debug(f"📅 [ROUTE] Sample row dates - due_date type: {type(due_date)}, value: {due_date}")
#
#         # Cache guard ~1MB
#         if use_cache and len(str(result.get('data', []))) < 1_000_000:
#             cache.set(report_name, filters, result, ttl=ttl)
#
#         # Format filter dates for output (no time) - ONLY FOR FILTERS, NOT DATA
#         result_filters = _format_filter_dates(result.get("filters", {}))
#         log.debug(f"🔧 [ROUTE] Formatted filters for output: {result_filters}")
#
#         result = {**result, "filters": result_filters}
#         resp = make_response(jsonify({
#             "success": True,
#             "cached": False,
#             **result
#         }))
#
#         # Response cache headers
#         if use_cache:
#             resp.headers["Cache-Control"] = f"public, max-age={ttl}"
#         else:
#             resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
#             resp.headers["Pragma"] = "no-cache"
#             resp.headers["Expires"] = "0"
#
#         return resp
#
#     except Forbidden as e:
#         return jsonify({"success": False, "error": str(e)}), 403
#     except ValueError as e:
#         return jsonify({"success": False, "error": str(e)}), 404
#     except PermissionError:
#         return jsonify({"success": False, "error": "Unauthorized"}), 401
#     except Exception as e:
#         log.error(f"Report execution failed: {e}", exc_info=True)
#         return jsonify({"success": False, "error": "Internal server error"}), 500
#
#
# @bp.route('/<report_name>/columns', methods=['GET'])
# def get_report_columns(report_name: str):
#     """Get column definitions for a report (no tenant data)."""
#     try:
#         engine = _get_engine()
#         filters = parse_filters(request)  # allow dynamic columns, no scope required
#         columns = engine.get_report_columns(report_name, filters)
#         return jsonify({
#             "success": True,
#             "report": report_name,
#             "columns": columns
#         })
#     except Exception as e:
#         return jsonify({"success": False, "error": str(e)}), 500
#
#
# @bp.route('/<report_name>/filters', methods=['GET'])
# def get_report_filters(report_name: str):
#     """Get filter definitions for a report (no tenant data)."""
#     try:
#         engine = _get_engine()
#         filters = engine.get_report_filters(report_name)
#         return jsonify({
#             "success": True,
#             "report": report_name,
#             "filters": filters
#         })
#     except Exception as e:
#         return jsonify({"success": False, "error": str(e)}), 500
#
#
# @bp.route('/<report_name>/meta', methods=['GET'])
# def get_report_meta(report_name: str):
#     """Get report metadata (no tenant data)."""
#     try:
#         engine = _get_engine()
#         meta = engine.get_report_meta(report_name)
#         return jsonify({
#             "success": True,
#             "report": report_name,
#             "meta": {
#                 "name": meta.name,
#                 "description": meta.description,
#                 "type": meta.report_type.value,
#                 "module": meta.module,
#                 "category": meta.category,
#                 "version": meta.version,
#                 "is_standard": meta.is_standard,
#                 "cache_enabled": meta.cache_enabled,
#                 "cache_ttl_s": meta.cache_ttl_s,
#             }
#         })
#     except Exception as e:
#         return jsonify({"success": False, "error": str(e)}), 500
#
#
# @bp.route('/list', methods=['GET'])
# def list_reports():
#     """List available reports (no tenant data)."""
#     try:
#         engine = _get_engine()
#         module = request.args.get('module')
#         category = request.args.get('category')
#         reports = engine.list_reports(module=module, category=category)
#         return jsonify({
#             "success": True,
#             "reports": reports,
#             "filters": {"module": module, "category": category}
#         })
#     except Exception as e:
#         return jsonify({"success": False, "error": str(e)}), 500
#
#
# @bp.route('/cache/clear', methods=['POST'])
# def clear_cache():
#     """
#     Clear / bump report cache.
#
#     Prefer: { "report_name": "...", "company_id": 2 } -> O(1) version bump.
#     Fallbacks (scan deletes) are provided for ops convenience.
#     """
#     try:
#         cache = _get_cache()
#         payload = request.get_json(silent=True) or {}
#         report_name = payload.get('report_name')
#         company_id = payload.get('company_id')
#
#         if report_name and company_id:
#             cache.bump_company(report_name, int(company_id))
#             message = f"Bumped version for {report_name} company={company_id}"
#             return jsonify({"success": True, "message": message})
#
#         if report_name:
#             cleared = cache.invalidate_report_all_companies(report_name)
#             return jsonify(
#                 {"success": True, "message": f"Cleared all companies for {report_name}", "cleared_keys": cleared})
#
#         if company_id:
#             cleared = cache.invalidate_company_all_reports(int(company_id))
#             return jsonify(
#                 {"success": True, "message": f"Cleared all reports for company {company_id}", "cleared_keys": cleared})
#
#         # no parameters: no-op
#         return jsonify({"success": True, "message": "Nothing to clear without parameters"})
#
#     except Exception as e:
#         return jsonify({"success": False, "error": str(e)}), 500
#
#
# @bp.route('/cache/stats', methods=['GET'])
# def cache_stats():
#     """Get cache statistics (no tenant data)."""
#     try:
#         cache = _get_cache()
#         stats = cache.get_stats()
#         return jsonify({"success": True, "stats": stats})
#     except Exception as e:
#         return jsonify({"success": False, "error": str(e)}), 500
#
#
# @bp.route('/export/<report_name>', methods=['POST'])
# def export_report(report_name: str):
#     """Export report to CSV/Excel (PDF placeholder). Requires company scope."""
#     try:
#         context = _ctx()
#         engine = _get_engine()
#
#         filters = request.get_json(silent=True) or {}
#         export_format = filters.pop('export_format', 'csv')
#
#         company_id = _to_int_or_none(filters.get('company'))
#         branch_id = _to_int_or_none(filters.get('branch_id'))
#         validate_report_access(context=context, company_id=company_id, branch_id=branch_id)
#
#         result = engine.execute_report(report_name, filters, context)
#
#         if export_format == 'csv':
#             return export_to_csv(result, report_name)
#         if export_format == 'excel':
#             return export_to_excel(result, report_name)
#         if export_format == 'pdf':
#             return export_to_pdf(result, report_name)
#         return jsonify({"success": False, "error": "Unsupported export format"}), 400
#
#     except Forbidden as e:
#         return jsonify({"success": False, "error": str(e)}), 403
#     except PermissionError:
#         return jsonify({"success": False, "error": "Unauthorized"}), 401
#     except Exception as e:
#         log.error(f"Report export failed: {e}", exc_info=True)
#         return jsonify({"success": False, "error": "Export failed"}), 500
#
#
# @bp.route('/document/<doc_type>/<doc_id>/<report_type>', methods=['GET'])
# def document_report(doc_type: str, doc_id: str, report_type: str):
#     """
#     Get report for specific document (scope enforced against the doc's company).
#     Example: GET /api/reports/document/PurchaseReceipt/123/general-ledger
#     """
#     try:
#         context = _ctx()
#         engine = _get_engine()
#
#         doc = get_document(doc_type, doc_id)
#         if not doc:
#             return jsonify({"success": False, "error": "Document not found"}), 404
#
#         company_id = getattr(doc, "company_id", None)
#         validate_report_access(context=context, company_id=company_id)
#
#         report_map = {
#             'general-ledger': 'General Ledger',
#             'stock-ledger': 'Stock Ledger',
#             'accounts-receivable': 'Accounts Receivable Summary',
#             'accounts-payable': 'Accounts Payable Summary'
#         }
#         if report_type not in report_map:
#             return jsonify({"success": False, "error": "Invalid report type"}), 400
#
#         report_name = report_map[report_type]
#
#         filters = {
#             'company': company_id,
#             'source_document': getattr(doc, "code", None),
#             'from_date': request.args.get('from_date'),
#             'to_date': request.args.get('to_date'),
#         }
#         filters = {k: v for k, v in filters.items() if v is not None}
#
#         result = engine.execute_report(report_name, filters, context)
#         # Format filter dates
#         result_filters = _format_filter_dates(result.get("filters", {}))
#
#         return jsonify({
#             "success": True,
#             "report": report_name,
#             "document_type": doc_type,
#             "document_id": doc_id,
#             "document_code": getattr(doc, "code", None),
#             "columns": result["columns"],
#             "data": result["data"],
#             "filters": result_filters
#         })
#
#     except Forbidden as e:
#         return jsonify({"success": False, "error": str(e)}), 403
#     except PermissionError:
#         return jsonify({"success": False, "error": "Unauthorized"}), 401
#     except Exception as e:
#         log.error(f"Document report failed: {e}", exc_info=True)
#         return jsonify({"success": False, "error": "Internal server error"}), 500
#
#
# # ---- helpers -----------------------------------------------------------------
# def parse_filters(request) -> Dict[str, Any]:
#     """
#     Parse and normalize filters from the request.
#     - Accept flexible date inputs; normalize to date objects.
#     - Accept bools (yes/true/1/on).
#     - Accept ints for known numeric fields.
#     - Includes 'nocache' and 'cache' flags.
#     """
#     body_filters = request.get_json(silent=True) or {}
#     query_filters = dict(request.args)
#     filters = {**body_filters, **query_filters} if request.method == "POST" else query_filters
#
#     log.debug(f"🔧 [PARSE_FILTERS] Raw filters: {filters}")
#
#     # Convert date strings (flex parser; no time)
#     date_fields = ['from_date', 'to_date', 'report_date', 'period_date', 'as_on_date', 'date']
#     for date_field in date_fields:
#         v = filters.get(date_field)
#         if v is not None:
#             if isinstance(v, str):
#                 parsed = _parse_date_flex(v)
#                 if parsed:
#                     filters[date_field] = parsed
#                     log.debug(f"🔧 [PARSE_FILTERS] Parsed {date_field}: '{v}' -> {parsed}")
#                 else:
#                     log.debug(f"🔧 [PARSE_FILTERS] Could not parse {date_field}: '{v}'")
#             elif isinstance(v, (date, datetime)):
#                 # Already a date object, ensure it's date not datetime
#                 if isinstance(v, datetime):
#                     filters[date_field] = v.date()
#                     log.debug(f"🔧 [PARSE_FILTERS] Converted datetime to date for {date_field}: {v.date()}")
#                 else:
#                     log.debug(f"🔧 [PARSE_FILTERS] {date_field} already date object: {v}")
#
#     # Convert boolean strings (including cache controls)
#     for bool_field in [
#         'show_zero_rows', 'show_unclosed_fy_pl', 'include_cancelled',
#         'include_provisional', 'compact', 'nocache', 'cache'
#     ]:
#         v = filters.get(bool_field)
#         if isinstance(v, str):
#             filters[bool_field] = v.lower() in ('true', '1', 'yes', 'y', 'on')
#             log.debug(f"🔧 [PARSE_FILTERS] Converted {bool_field}: '{v}' -> {filters[bool_field]}")
#
#     # Convert integer-like strings
#     for num_field in [
#         'range1', 'range2', 'range3', 'range4',
#         'company', 'branch_id', 'party',
#         'item_id', 'warehouse_id', 'limit', 'offset'
#     ]:
#         v = filters.get(num_field)
#         try:
#             if isinstance(v, str) and v.strip() != "":
#                 filters[num_field] = int(v)
#                 log.debug(f"🔧 [PARSE_FILTERS] Converted {num_field}: '{v}' -> {filters[num_field]}")
#         except Exception as e:
#             log.debug(f"🔧 [PARSE_FILTERS] Failed to convert {num_field}: '{v}' - {e}")
#
#     log.debug(f"🔧 [PARSE_FILTERS] Final parsed filters: {filters}")
#     return filters
#
#
# def export_to_csv(result: Dict[str, Any], report_name: str) -> Response:
#     output = io.StringIO()
#     writer = csv.writer(output)
#     headers = [col['label'] for col in result['columns']]
#     writer.writerow(headers)
#     for row in result['data']:
#         writer.writerow([row.get(col['fieldname'], '') for col in result['columns']])
#     filename = f"{report_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
#     return Response(output.getvalue(), mimetype="text/csv",
#                     headers={"Content-Disposition": f"attachment;filename={filename}"})
#
#
# def export_to_excel(result: Dict[str, Any], report_name: str) -> Response:
#     try:
#         import openpyxl
#         from openpyxl.utils import get_column_letter
#
#         workbook = openpyxl.Workbook()
#         worksheet = workbook.active
#         worksheet.title = report_name[:31]
#
#         for col_idx, column in enumerate(result['columns'], 1):
#             cell = worksheet.cell(row=1, column=col_idx)
#             cell.value = column['label']
#             cell.font = openpyxl.styles.Font(bold=True)
#
#         for row_idx, row_data in enumerate(result['data'], 2):
#             for col_idx, column in enumerate(result['columns'], 1):
#                 worksheet.cell(row=row_idx, column=col_idx).value = row_data.get(column['fieldname'], '')
#
#         for column_cells in worksheet.columns:
#             length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
#             worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
#
#         output = io.BytesIO()
#         workbook.save(output)
#         output.seek(0)
#
#         filename = f"{report_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
#         return Response(output.getvalue(),
#                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                         headers={"Content-Disposition": f"attachment;filename={filename}"})
#     except ImportError:
#         return jsonify({"success": False, "error": "Excel export requires openpyxl"}), 500
#
#
# def export_to_pdf(result: Dict[str, Any], report_name: str) -> Response:
#     return jsonify({"success": False, "error": "PDF export not yet implemented"}), 501
#
#
# def get_document(doc_type: str, doc_id: str):
#     try:
#         from app.application_stock.stock_models import StockEntry
#         from app.application_buying.models import SalesInvoice, PurchaseInvoice, PurchaseReceipt
#
#         model_map = {
#             'PurchaseReceipt': PurchaseReceipt,
#             'StockEntry': StockEntry,
#             'SalesInvoice': SalesInvoice,
#             'PurchaseInvoice': PurchaseInvoice,
#         }
#         Model = model_map.get(doc_type)
#         return Model.query.get(doc_id) if Model else None
#     except Exception as e:
#         log.error(f"Error getting document {doc_type} {doc_id}: {e}")
#         return None
# app/application_reports/routes.py
from __future__ import annotations
import logging
import csv
import io
from typing import Dict, Any
from datetime import datetime, date

from flask import Blueprint, request, jsonify, current_app, Response, g, make_response
from werkzeug.exceptions import Forbidden

from app.application_reports.core.engine import create_report_engine
from app.application_reports.bootstrap import bootstrap_reports
from app.application_reports.core.cache import ReportCache
from app.application_reports.security import validate_report_access
from app.auth.deps import get_current_user
from app.security.rbac_effective import AffiliationContext
from config.database import db

bp = Blueprint('reports', __name__, url_prefix='/api/reports')
log = logging.getLogger(__name__)

_HOT_REPORTS = {
    "General Ledger",
    "Accounts Payable Summary",
    "Accounts Receivable Summary",
    "Accounts Payable",
    "Accounts Receivable",
    "Stock Ledger",
    "Trial Balance",
    "Balance Sheet",
    "Profit and Loss",
}

def _ctx() -> AffiliationContext:
    _ = get_current_user()
    ctx: AffiliationContext = getattr(g, "auth", None)
    if not ctx:
        raise PermissionError("Authentication context not found.")
    return ctx

_engine_instance = None
_report_cache: ReportCache | None = None

def _get_engine():
    global _engine_instance
    if _engine_instance is None:
        _engine_instance = create_report_engine(db.session)
        bootstrap_reports(_engine_instance)
        log.info("✅ Reports engine initialized and bootstrapped")
    return _engine_instance

def _get_cache() -> ReportCache:
    global _report_cache
    if _report_cache is None:
        enabled = current_app.config.get('REPORT_CACHE_ENABLED', True)
        ttl = int(current_app.config.get('REPORT_CACHE_TTL', 30))
        _report_cache = ReportCache(enabled=enabled, default_ttl=ttl)
        log.info("🧠 Report cache initialized (enabled=%s, ttl=%s)", enabled, ttl)
    return _report_cache

def _to_int_or_none(v):
    try:
        return int(v) if v is not None and str(v).strip() != "" else None
    except Exception:
        return None

_DISPLAY_FMT = "%d-%m-%Y"
_DATE_PARSE_PATTERNS = [
    "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d",
    "%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y",
]

def _parse_date_flex(s: str) -> date | None:
    s = (s or "").strip()
    if not s:
        return None
    if "T" in s or " " in s:
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            return dt.date()
        except Exception:
            pass
    for fmt in _DATE_PARSE_PATTERNS:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None

def _format_date_out(d: date | datetime | None) -> str | None:
    if d is None:
        return None
    if isinstance(d, datetime):
        d = d.date()
    return d.strftime(_DISPLAY_FMT)

def _format_filter_dates(filters: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(filters, dict):
        return filters
    out = dict(filters)
    date_keys = set(k for k in out.keys()
                    if k.endswith("_date") or k in {"date", "as_on_date", "period_date", "report_date"})
    for k in date_keys:
        v = out.get(k)
        if isinstance(v, (date, datetime)):
            out[k] = _format_date_out(v)
        elif isinstance(v, str):
            parsed = _parse_date_flex(v)
            out[k] = _format_date_out(parsed) if parsed else v
    return out

def _booly(v) -> bool | None:
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.lower() in ("true", "1", "yes", "y", "on")
    return None

def _should_cache(report_name: str, report_meta, filters: Dict[str, Any]) -> tuple[bool, int]:
    client_nocache = _booly(filters.get('nocache'))
    client_cache_flag = _booly(filters.get('cache'))
    if client_nocache or (client_cache_flag is False):
        return (False, 0)
    if report_meta and report_meta.cache_enabled is not None:
        enabled = bool(report_meta.cache_enabled)
    else:
        enabled = (report_name not in _HOT_REPORTS)
    ttl = int(report_meta.cache_ttl_s or current_app.config.get('REPORT_CACHE_TTL', 30))
    return (enabled, ttl)

@bp.route('/<report_name>', methods=['GET', 'POST'])
def execute_report(report_name: str):
    try:
        context = _ctx()
        engine = _get_engine()

        filters = parse_filters(request)
        log.debug(f"🔧 [ROUTE] Initial filters for {report_name}: {filters}")

        company_id = _to_int_or_none(filters.get('company'))
        branch_id = _to_int_or_none(filters.get('branch_id'))
        validate_report_access(context=context, company_id=company_id, branch_id=branch_id)

        report_meta = engine.get_report_meta(report_name)
        use_cache, ttl = _should_cache(report_name, report_meta, filters)

        cache = _get_cache()
        if use_cache:
            cached_result = cache.get(report_name, filters)
            if cached_result:
                cached_result = {**cached_result, "filters": _format_filter_dates(cached_result.get("filters", {}))}
                resp = make_response(jsonify({
                    "success": True,
                    "cached": True,
                    **cached_result
                }))
                resp.headers["Cache-Control"] = f"public, max-age={ttl}"
                return resp

        log.debug(f"🚀 [ROUTE] Executing report {report_name} with filters: {filters}")
        result = engine.execute_report(report_name, filters, context)
        log.debug(f"✅ [ROUTE] Report executed successfully. Data rows: {len(result.get('data', []))}")

        if use_cache and len(str(result.get('data', []))) < 1_000_000:
            cache.set(report_name, filters, result, ttl=ttl)

        result_filters = _format_filter_dates(result.get("filters", {}))
        log.debug(f"🔧 [ROUTE] Formatted filters for output: {result_filters}")

        result = {**result, "filters": result_filters}
        resp = make_response(jsonify({
            "success": True,
            "cached": False,
            **result
        }))

        if use_cache:
            resp.headers["Cache-Control"] = f"public, max-age={ttl}"
        else:
            resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            resp.headers["Pragma"] = "no-cache"
            resp.headers["Expires"] = "0"

        return resp

    except Forbidden as e:
        return jsonify({"success": False, "error": str(e)}), 403
    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 404
    except PermissionError:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    except Exception as e:
        log.error(f"Report execution failed: {e}", exc_info=True)
        return jsonify({"success": False, "error": "Internal server error"}), 500


@bp.route('/<report_name>/columns', methods=['GET'])
def get_report_columns(report_name: str):
    try:
        engine = _get_engine()
        filters = parse_filters(request)
        columns = engine.get_report_columns(report_name, filters)
        return jsonify({
            "success": True,
            "report": report_name,
            "columns": columns
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route('/<report_name>/filters', methods=['GET'])
def get_report_filters(report_name: str):
    try:
        engine = _get_engine()
        filters = engine.get_report_filters(report_name)
        return jsonify({
            "success": True,
            "report": report_name,
            "filters": filters
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route('/<report_name>/meta', methods=['GET'])
def get_report_meta(report_name: str):
    try:
        engine = _get_engine()
        meta = engine.get_report_meta(report_name)
        return jsonify({
            "success": True,
            "report": report_name,
            "meta": {
                "name": meta.name,
                "description": meta.description,
                "type": meta.report_type.value,
                "module": meta.module,
                "category": meta.category,
                "version": meta.version,
                "is_standard": meta.is_standard,
                "cache_enabled": meta.cache_enabled,
                "cache_ttl_s": meta.cache_ttl_s,
            }
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route('/list', methods=['GET'])
def list_reports():
    try:
        engine = _get_engine()
        module = request.args.get('module')
        category = request.args.get('category')
        reports = engine.list_reports(module=module, category=category)
        return jsonify({
            "success": True,
            "reports": reports,
            "filters": {"module": module, "category": category}
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route('/cache/clear', methods=['POST'])
def clear_cache():
    try:
        cache = _get_cache()
        payload = request.get_json(silent=True) or {}
        report_name = payload.get('report_name')
        company_id = payload.get('company_id')

        if report_name and company_id:
            cache.bump_company(report_name, int(company_id))
            message = f"Bumped version for {report_name} company={company_id}"
            return jsonify({"success": True, "message": message})

        if report_name:
            # NOTE: if you have a helper invalidate_report_all_companies, use it here.
            cleared = cache.invalidate_report(report_name)
            return jsonify({"success": True, "message": f"Cleared all companies for {report_name}", "cleared_keys": cleared})

        if company_id:
            cleared = cache.invalidate_company_all_reports(int(company_id))
            return jsonify({"success": True, "message": f"Cleared all reports for company {company_id}", "cleared_keys": cleared})

        return jsonify({"success": True, "message": "Nothing to clear without parameters"})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route('/cache/stats', methods=['GET'])
def cache_stats():
    try:
        cache = _get_cache()
        stats = cache.get_stats()
        return jsonify({"success": True, "stats": stats})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route('/export/<report_name>', methods=['POST'])
def export_report(report_name: str):
    try:
        context = _ctx()
        engine = _get_engine()

        filters = request.get_json(silent=True) or {}
        export_format = filters.pop('export_format', 'csv')

        company_id = _to_int_or_none(filters.get('company'))
        branch_id = _to_int_or_none(filters.get('branch_id'))
        validate_report_access(context=context, company_id=company_id, branch_id=branch_id)

        result = engine.execute_report(report_name, filters, context)

        if export_format == 'csv':
            return export_to_csv(result, report_name)
        if export_format == 'excel':
            return export_to_excel(result, report_name)
        if export_format == 'pdf':
            return export_to_pdf(result, report_name)
        return jsonify({"success": False, "error": "Unsupported export format"}), 400

    except Forbidden as e:
        return jsonify({"success": False, "error": str(e)}), 403
    except PermissionError:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    except Exception as e:
        log.error(f"Report export failed: {e}", exc_info=True)
        return jsonify({"success": False, "error": "Export failed"}), 500


@bp.route('/document/<doc_type>/<doc_id>/<report_type>', methods=['GET'])
def document_report(doc_type: str, doc_id: str, report_type: str):
    try:
        context = _ctx()
        engine = _get_engine()

        doc = get_document(doc_type, doc_id)
        if not doc:
            return jsonify({"success": False, "error": "Document not found"}), 404

        company_id = getattr(doc, "company_id", None)
        validate_report_access(context=context, company_id=company_id)

        report_map = {
            'general-ledger': 'General Ledger',
            'stock-ledger': 'Stock Ledger',
            'accounts-receivable': 'Accounts Receivable Summary',
            'accounts-payable': 'Accounts Payable Summary'
        }
        if report_type not in report_map:
            return jsonify({"success": False, "error": "Invalid report type"}), 400

        report_name = report_map[report_type]

        filters = {
            'company': company_id,
            'source_document': getattr(doc, "code", None),
            'from_date': request.args.get('from_date'),
            'to_date': request.args.get('to_date'),
        }
        filters = {k: v for k, v in filters.items() if v is not None}

        result = engine.execute_report(report_name, filters, context)
        result_filters = _format_filter_dates(result.get("filters", {}))

        return jsonify({
            "success": True,
            "report": report_name,
            "document_type": doc_type,
            "document_id": doc_id,
            "document_code": getattr(doc, "code", None),
            "columns": result["columns"],
            "data": result["data"],
            "filters": result_filters
        })

    except Forbidden as e:
        return jsonify({"success": False, "error": str(e)}), 403
    except PermissionError:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    except Exception as e:
        log.error(f"Document report failed: {e}", exc_info=True)
        return jsonify({"success": False, "error": "Internal server error"}), 500


def parse_filters(request) -> Dict[str, Any]:
    body_filters = request.get_json(silent=True) or {}
    query_filters = dict(request.args)
    filters = {**body_filters, **query_filters} if request.method == "POST" else query_filters

    log.debug(f"🔧 [PARSE_FILTERS] Raw filters: {filters}")

    date_fields = ['from_date', 'to_date', 'report_date', 'period_date', 'as_on_date', 'date']
    for date_field in date_fields:
        v = filters.get(date_field)
        if v is not None:
            if isinstance(v, str):
                parsed = _parse_date_flex(v)
                if parsed:
                    filters[date_field] = parsed
                    log.debug(f"🔧 [PARSE_FILTERS] Parsed {date_field}: '{v}' -> {parsed}")
            elif isinstance(v, (date, datetime)):
                if isinstance(v, datetime):
                    filters[date_field] = v.date()
                    log.debug(f"🔧 [PARSE_FILTERS] Converted datetime to date for {date_field}: {v.date()}")

    for bool_field in [
        'show_zero_rows', 'show_unclosed_fy_pl', 'include_cancelled',
        'include_provisional', 'compact', 'nocache', 'cache'
    ]:
        v = filters.get(bool_field)
        if isinstance(v, str):
            filters[bool_field] = v.lower() in ('true', '1', 'yes', 'y', 'on')
            log.debug(f"🔧 [PARSE_FILTERS] Converted {bool_field}: '{v}' -> {filters[bool_field]}")

    for num_field in [
        'range1', 'range2', 'range3', 'range4',
        'company', 'branch_id', 'party',
        'item_id', 'warehouse_id', 'limit', 'offset', 'page_length', 'start'
    ]:
        v = filters.get(num_field)
        try:
            if isinstance(v, str) and v.strip() != "":
                filters[num_field] = int(v)
                log.debug(f"🔧 [PARSE_FILTERS] Converted {num_field}: '{v}' -> {filters[num_field]}")
        except Exception as e:
            log.debug(f"🔧 [PARSE_FILTERS] Failed to convert {num_field}: '{v}' - {e}")

    log.debug(f"🔧 [PARSE_FILTERS] Final parsed filters: {filters}")
    return filters


def export_to_csv(result: Dict[str, Any], report_name: str) -> Response:
    output = io.StringIO()
    writer = csv.writer(output)
    headers = [col['label'] for col in result['columns']]
    writer.writerow(headers)
    for row in result['data']:
        writer.writerow([row.get(col['fieldname'], '') for col in result['columns']])
    filename = f"{report_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment;filename={filename}"})


def export_to_excel(result: Dict[str, Any], report_name: str) -> Response:
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = report_name[:31]

        for col_idx, column in enumerate(result['columns'], 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = column['label']
            cell.font = openpyxl.styles.Font(bold=True)

        for row_idx, row_data in enumerate(result['data'], 2):
            for col_idx, column in enumerate(result['columns'], 1):
                worksheet.cell(row=row_idx, column=col_idx).value = row_data.get(column['fieldname'], '')

        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"{report_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(output.getvalue(),
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment;filename={filename}"})
    except ImportError:
        return jsonify({"success": False, "error": "Excel export requires openpyxl"}), 500


def export_to_pdf(result: Dict[str, Any], report_name: str) -> Response:
    return jsonify({"success": False, "error": "PDF export not yet implemented"}), 501


def get_document(doc_type: str, doc_id: str):
    try:
        from app.application_stock.stock_models import StockEntry
        from app.application_buying.models import SalesInvoice, PurchaseInvoice, PurchaseReceipt

        model_map = {
            'PurchaseReceipt': PurchaseReceipt,
            'StockEntry': StockEntry,
            'SalesInvoice': SalesInvoice,
            'PurchaseInvoice': PurchaseInvoice,
        }
        Model = model_map.get(doc_type)
        return Model.query.get(doc_id) if Model else None
    except Exception as e:
        log.error(f"Error getting document {doc_type} {doc_id}: {e}")
        return None
