# application_data_import/parsers/xlsx_reader.py
from __future__ import annotations
from typing import List, Dict, Any
import io


def read_xlsx_bytes(raw: bytes) -> (List[str], List[Dict[str, Any]]):
    from openpyxl import load_workbook
    f = io.BytesIO(raw)
    wb = load_workbook(filename=f, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(c or "").strip() for c in rows[0]]
    out_rows: List[Dict[str, Any]] = []
    for r in rows[1:]:
        out_rows.append({headers[i]: r[i] for i in range(len(headers))})
    return headers, out_rows
